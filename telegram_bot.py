#!/usr/bin/env python3
"""Telegram bot front-end (python-telegram-bot v21, async polling).

Heavy work (LLM, VLM, history I/O) runs in asyncio.to_thread.
tg_push() at the bottom is a sync helper for cron scripts that need to send
messages without an event loop (wakeup.py, reminder_check.py, inbox_scan.py).
"""

from __future__ import annotations

import asyncio
import os
import re
import sys
import time
import pathlib
import urllib.parse
import urllib.request

sys.path.insert(0, str(Path(__file__).resolve().parent))
import core
import media
import memory as memory_module

from telegram import Update
from telegram.constants import ChatAction
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

BUBBLE_DELAY = 0.5
MAX_BUBBLE_CHARS = 3500
MAX_FOLLOWUP_BUBBLES = 60

TG_BOT_TOKEN = os.environ.get("TG_BOT_TOKEN", "")
TG_DEFAULT_CHAT_ID = os.environ.get("TG_DEFAULT_CHAT_ID", "")
TG_PROXY = os.environ.get("TG_PROXY", "") or os.environ.get("HTTPS_PROXY", "")
HEARTBEAT_FILE = pathlib.Path(__file__).resolve().parent / "poll_heartbeat"


def _parse_allowed() -> set[int]:
    raw = os.environ.get("TG_ALLOWED_CHAT_IDS", "")
    out: set[int] = set()
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            out.add(int(part))
        except ValueError:
            core.log(f"tg: bad chat_id in TG_ALLOWED_CHAT_IDS: {part}")
    return out


ALLOWED_CHAT_IDS = _parse_allowed()

# Markers that should never reach the user as raw text.
KNOWN_MARKERS = (
    "FETCH", "BROWSE", "POI", "RECALL", "INBOX", "TOOL", "RESEARCH",
    "EMAIL_START", "EMAIL_END", "EMAIL", "YT", "DRAW",
    "MEMO", "SELF", "DIARY", "PREF", "voice",
)
_LEAK_PATTERN = re.compile(
    r"\[(?:" + "|".join(KNOWN_MARKERS) + r")(?::[^\]]*|\b[^\]]*)\]",
    re.IGNORECASE | re.DOTALL,
)

# Phrases that explicitly ask for audio output.
# Used as a code-level fallback when the AI omits [voice: ...] markers.
_VOICE_REQUEST_RE = re.compile(
    r"念一下|念给我|念出来|读一下|读给我|读出来|朗读|大声读|大声念"
    r"|用语音|发语音|语音说|发条语音|read aloud|say aloud",
    re.IGNORECASE,
)

WELCOME = (
    "你好，我是 soy \U0001f916\n\n"
    "直接发消息聊天，也能看图（发图片+caption）、读文件（PDF/docx/txt等）。\n\n"
    "命令：\n"
    "/clear   清空对话历史\n"
    "/status  后台任务状态\n"
    "/cost    费用统计\n"
    "/ds <内容>  强制走 DeepSeek\n"
    "/cf <内容>  强制走 Claude(CF)\n"
    "/mm <内容>  强制走 MiniMax\n"
    "/q  <内容>  Qwen3 工具调用（搜索/调研/邮件）\n"
)


def _is_allowed(chat_id: int) -> bool:
    return not ALLOWED_CHAT_IDS or chat_id in ALLOWED_CHAT_IDS


def _chat_key(chat_id: int) -> str:
    return f"tg_{chat_id}"


# ---------------------------------------------------------------------------
# Text utilities
# ---------------------------------------------------------------------------


def _chunk_long(text: str, limit: int) -> list[str]:
    """Split a single string that exceeds `limit` chars at natural break points."""
    if len(text) <= limit:
        return [text]
    chunks: list[str] = []
    while len(text) > limit:
        cut = -1
        for delim in ("。", "！", "？", "\n", ". ", " "):
            pos = text.rfind(delim, 0, limit)
            if pos > 0:
                cut = pos + len(delim)
                break
        if cut <= 0:
            cut = limit
        chunks.append(text[:cut].strip())
        text = text[cut:].strip()
    if text:
        chunks.append(text)
    return [c for c in chunks if c]


def split_bubbles(text: str) -> list[str]:
    """Split a reply into Telegram-style chat bubbles.

    Steps: sentence-split single paragraphs → filter decoration lines →
    merge only very short fragments (< 12 chars) → chunk anything too long.
    """
    text = core.auto_split_sentences(text)
    raw: list[str] = []
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue
        if re.fullmatch(r"[-_=*#~`|><.]+", line):
            continue
        if not re.search(r"[\w一-鿿]", line):
            continue
        raw.append(line)
    bubbles: list[str] = []
    for line in raw:
        if (
            bubbles
            and len(bubbles[-1]) < 12
            and not re.match(r"^[\d\-\*•－#]", bubbles[-1])
            and not re.match(r"^[\d\-\*•－#]", line)
        ):
            bubbles[-1] = bubbles[-1] + " " + line
        else:
            bubbles.append(line)
    out: list[str] = []
    for b in bubbles:
        out.extend(_chunk_long(b, MAX_BUBBLE_CHARS))
    return out


def strip_markers(reply: str, user_id: str) -> str:
    """Run storage markers (MEMO/SELF/DIARY), replace [voice:] with plain text,
    strip all other known markers so the DSL never leaks to the user."""
    reply = core.process_markers(reply, user_id=user_id)
    reply = re.sub(r"\[voice:\s*(.*?)\]", r"\1", reply, flags=re.DOTALL)
    reply = _LEAK_PATTERN.sub("", reply)
    return re.sub(r"\n{3,}", "\n\n", reply).strip()


# ---------------------------------------------------------------------------
# Send helpers
# ---------------------------------------------------------------------------


async def _send_bubbles(update: Update, bubbles: list[str]) -> None:
    for b in bubbles[:MAX_FOLLOWUP_BUBBLES]:
        if not b.strip():
            continue
        try:
            await update.effective_chat.send_message(b)
        except Exception as e:
            core.log(f"tg send-bubble: {e}")
        await asyncio.sleep(BUBBLE_DELAY)


async def _reject_silent(update: Update) -> bool:
    chat = update.effective_chat
    if chat is None:
        return True
    if not _is_allowed(chat.id):
        core.log(f"tg: rejected chat_id={chat.id}")
        return True
    return False


# ---------------------------------------------------------------------------
# Command handlers
# ---------------------------------------------------------------------------


async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    if await _reject_silent(update):
        return
    await update.message.reply_text(WELCOME)


async def cmd_clear(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    if await _reject_silent(update):
        return
    key = _chat_key(update.effective_chat.id)
    await asyncio.to_thread(memory_module.save_history, key, [])
    await update.message.reply_text("History cleared.")


async def cmd_status(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    if await _reject_silent(update):
        return
    text = await asyncio.to_thread(_build_status)
    await update.message.reply_text(text)


async def cmd_cost(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    if await _reject_silent(update):
        return
    key = _chat_key(update.effective_chat.id)
    text = await asyncio.to_thread(_build_cost, key)
    await update.message.reply_text(text)


def _build_status() -> str:
    lines: list[str] = []
    now = time.time()

    research = core.list_research_tasks()
    if research:
        lines.append(f"## Background research ({len(research)})")
        for r in research:
            tag = "\U0001f4e7+tg" if r["email"] else "tg"
            lines.append(f"- {r['id']} | {tag} | {r['elapsed']}s | {r['task'][:60]}")
    else:
        lines.append("## Background research: none")

    try:
        rows = memory_module.supa_select(
            "reminders",
            {"select": "id,time,date,content", "done": "eq.false", "order": "id.asc", "limit": "100"},
        )
        if rows:
            lines.append(f"## Pending reminders: {len(rows)}")
            for r in rows[:5]:
                lines.append(f"- {r.get('date','?')} {r.get('time','?')}: {(r.get('content') or '')[:40]}")
            if len(rows) > 5:
                lines.append(f"  …and {len(rows) - 5} more")
        else:
            lines.append("## Pending reminders: 0")
    except Exception as e:
        lines.append(f"## Pending reminders: (err {str(e)[:40]})")

    cron_paths = {
        "wakeup": str(Path(__file__).resolve().parent / "wakeup.log"),
        "watchdog": str(Path(__file__).resolve().parent / "watchdog.log"),
        "reminder_check": str(Path(__file__).resolve().parent / "reminder.log"),
        "inbox_scan": str(Path(__file__).resolve().parent / "inbox_scan.log"),
        "consolidate": str(Path(__file__).resolve().parent / "consolidate.log"),
    }
    lines.append("## Recent cron")
    for name, path in cron_paths.items():
        try:
            elapsed = int(now - os.path.getmtime(path))
            if elapsed < 60:
                t = f"{elapsed}s ago"
            elif elapsed < 3600:
                t = f"{elapsed // 60}m ago"
            elif elapsed < 86400:
                t = f"{elapsed // 3600}h ago"
            else:
                t = f"{elapsed // 86400}d ago"
            lines.append(f"- {name}: {t}")
        except Exception:
            lines.append(f"- {name}: (no log yet)")
    return "\n".join(lines)


def _build_cost(chat_key: str) -> str:
    rows = memory_module.supa_select(
        "costs",
        {"select": "cost_usd,model,tool_type,ts", "user_id": f"eq.{chat_key}", "order": "ts.desc", "limit": "500"},
    )
    if not rows:
        return "No cost data yet."
    total = 0.0
    month_total = 0.0
    by_tool: dict[str, float] = {}
    month = time.strftime("%Y-%m")
    for row in rows:
        cost = float(row.get("cost_usd") or 0)
        total += cost
        tool = row.get("tool_type") or "unknown"
        by_tool[tool] = by_tool.get(tool, 0.0) + cost
        if str(row.get("ts") or "").startswith(month):
            month_total += cost
    lines = [f"month_total_usd: {month_total:.6f}", f"all_time_total_usd: {total:.6f}"]
    top = sorted(by_tool.items(), key=lambda x: x[1], reverse=True)[:3]
    if top:
        lines.append("top_tools:")
        lines.extend(f"- {tool}: {cost:.6f}" for tool, cost in top)
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Model-override command handlers (/ds /cf /mm /q)
# ---------------------------------------------------------------------------


async def _cmd_model(update: Update, ctx: ContextTypes.DEFAULT_TYPE, model: str) -> None:
    if await _reject_silent(update):
        return
    text = " ".join(ctx.args or []).strip()
    if not text:
        await update.message.reply_text(f"用法: /{model} <内容>")
        return
    chat_id_int = update.effective_chat.id
    key = _chat_key(chat_id_int)
    try:
        await ctx.bot.send_chat_action(chat_id_int, ChatAction.TYPING)
    except Exception:
        pass

    def _do() -> str:
        sys_p = core.get_system_prompt(key)
        msgs = core.build_messages(key, text)
        flat = core.build_prompt(key, text)
        if model == "ds":
            reply = core.call_deepseek(msgs, sys_p) or core.get_reply(key, text)
        elif model == "cf":
            reply = core.call_claude(flat) or core.get_reply(key, text)
        elif model == "mm":
            reply = core.call_minimax(msgs, sys_p) or core.get_reply(key, text)
        elif model == "q":
            reply = core.call_qwen(msgs, sys_p) or core.get_reply(key, text)
        else:
            reply = core.get_reply(key, text)
        if not reply:
            return "(empty)"
        cleaned = strip_markers(reply, user_id=key)
        if cleaned:
            core.append_history(key, "user", text, user_id=key)
            core.append_history(key, "assistant", cleaned, user_id=key)
        return cleaned or "(empty)"

    try:
        reply = await asyncio.to_thread(_do)
    except Exception as e:
        core.log(f"tg _cmd_model {model}: {e}")
        await update.message.reply_text(f"(error: {str(e)[:80]})")
        return

    bubbles = split_bubbles(reply) or [reply]
    await update.message.reply_text(bubbles[0])
    if len(bubbles) > 1:
        await _send_bubbles(update, bubbles[1:])


async def cmd_ds(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    await _cmd_model(update, ctx, "ds")


async def cmd_cf(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    await _cmd_model(update, ctx, "cf")


async def cmd_mm(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    await _cmd_model(update, ctx, "mm")


async def cmd_q(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    """Qwen3 tool-calling agent — good for search / research / email tasks."""
    if await _reject_silent(update):
        return
    text = " ".join(ctx.args or []).strip()
    if not text:
        await update.message.reply_text("用法: /q <调研/搜索/任务描述>")
        return
    chat_id_int = update.effective_chat.id
    key = _chat_key(chat_id_int)
    await update.message.reply_text("好，我去查，查完告诉你。")

    def _do() -> str:
        sys_p = core.get_system_prompt(key)
        msgs = core.build_messages(key, text)
        result = core.call_qwen(msgs, sys_p)
        if not result:
            return ""
        cleaned = strip_markers(result, user_id=key)
        if cleaned:
            core.append_history(key, "user", text, user_id=key)
            core.append_history(key, "assistant", cleaned, user_id=key)
        return cleaned or ""

    try:
        result = await asyncio.to_thread(_do)
    except Exception as e:
        core.log(f"tg cmd_q: {e}")
        await update.message.reply_text(f"(error: {str(e)[:80]})")
        return

    if not result:
        await update.message.reply_text("(Qwen 没有返回结果)")
        return

    bubbles = split_bubbles(result) or [result]
    for b in bubbles[:MAX_FOLLOWUP_BUBBLES]:
        if b.strip():
            await update.effective_chat.send_message(b)
            await asyncio.sleep(BUBBLE_DELAY)


# ---------------------------------------------------------------------------
# Message handlers
# ---------------------------------------------------------------------------


async def on_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    if await _reject_silent(update):
        return
    msg = update.message
    if msg is None or not msg.text:
        return
    chat_id_int = update.effective_chat.id
    key = _chat_key(chat_id_int)
    user_msg = msg.text.strip()
    if not user_msg:
        return
    core.log(f"<< tg:{chat_id_int}: {user_msg[:60]}")
    try:
        await ctx.bot.send_chat_action(chat_id_int, ChatAction.TYPING)
    except Exception:
        pass

    def _do_reply() -> tuple[list[tuple[str, str]], str]:
        raw = core.get_reply(key, user_msg) or ""
        processed = core.process_markers(raw, user_id=key)

        # Parse [voice: ...] segments before stripping them
        segs = core.parse_voice_segments(processed)

        # Strip leak markers from text segments (voice segments are kept as-is)
        display_segs: list[tuple[str, str]] = []
        for kind, content in segs:
            if kind == "text":
                content = _LEAK_PATTERN.sub("", content)
                content = re.sub(r"\n{3,}", "\n\n", content).strip()
            display_segs.append((kind, content))
        display_segs = [(k, c) for k, c in display_segs if c.strip()]

        # Cleaned text for history: voice marker content kept as plain text
        cleaned = re.sub(r"\[voice:\s*(.*?)\]", r"\1", processed, flags=re.DOTALL)
        cleaned = _LEAK_PATTERN.sub("", cleaned)
        cleaned = re.sub(r"\n{3,}", "\n\n", cleaned).strip()
        if cleaned:
            core.append_history(key, "user", user_msg, user_id=key)
            core.append_history(key, "assistant", cleaned, user_id=key)

        # Code-level fallback: user asked for audio but AI skipped [voice: ...]
        if not any(k == "voice" for k, _ in display_segs) and _VOICE_REQUEST_RE.search(user_msg):
            if cleaned:
                display_segs = [("voice", cleaned)]
                core.log(f"tg voice-force: {len(cleaned)}c")

        core.log(f">> tg:{chat_id_int}: {len(display_segs)} seg(s), cleaned={len(cleaned)}c")
        return display_segs, cleaned

    try:
        segs, cleaned = await asyncio.to_thread(_do_reply)
    except Exception as e:
        core.log(f"tg on_text: {e}")
        await msg.reply_text(f"(internal error: {str(e)[:80]})")
        return

    if not any(c.strip() for _, c in segs) and not cleaned:
        await msg.reply_text("(empty)")
        return

    first_sent = False
    for kind, content in segs:
        content = content.strip()
        if not content:
            continue
        if kind == "text":
            bubbles = split_bubbles(content) or [content]
            for b in bubbles[:MAX_FOLLOWUP_BUBBLES]:
                if not b.strip():
                    continue
                try:
                    if not first_sent:
                        await msg.reply_text(b)
                        first_sent = True
                    else:
                        await update.effective_chat.send_message(b)
                except Exception as e:
                    core.log(f"tg send bubble: {e}")
                await asyncio.sleep(BUBBLE_DELAY)
        else:
            # voice segment → TTS
            try:
                mp3 = await asyncio.to_thread(core.generate_tts, content)
                if mp3:
                    core.log(f"tg tts: ok {len(mp3)}b")
                    await update.effective_message.reply_voice(voice=mp3)
                    first_sent = True
                else:
                    core.log(f"tg tts: None for {len(content)}c, sending as text")
                    target = msg if not first_sent else update.effective_chat
                    await (target.reply_text(content) if not first_sent else target.send_message(content))
                    first_sent = True
            except Exception as e:
                core.log(f"tg tts error: {e}")
                try:
                    await update.effective_chat.send_message(content)
                    first_sent = True
                except Exception:
                    pass
            await asyncio.sleep(BUBBLE_DELAY)

    # Fallback: if all segs were empty but we have cleaned text
    if not first_sent and cleaned:
        bubbles = split_bubbles(cleaned) or [cleaned]
        await msg.reply_text(bubbles[0])
        if len(bubbles) > 1:
            await _send_bubbles(update, bubbles[1:])


async def on_photo(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    if await _reject_silent(update):
        return
    msg = update.message
    if not msg or not msg.photo:
        return
    chat_id_int = update.effective_chat.id
    key = _chat_key(chat_id_int)
    caption = (msg.caption or "").strip()
    core.log(f"<< tg:{chat_id_int}: [photo] caption={caption[:60]}")
    try:
        await ctx.bot.send_chat_action(chat_id_int, ChatAction.TYPING)
    except Exception:
        pass

    try:
        f = await ctx.bot.get_file(msg.photo[-1].file_id)
        image_bytes = bytes(await f.download_as_bytearray())
    except Exception as e:
        core.log(f"tg photo fetch: {e}")
        await msg.reply_text(f"(图片下载失败: {str(e)[:60]})")
        return

    # VLM → get visual description
    try:
        desc = await asyncio.to_thread(media.call_vlm, image_bytes, caption)
        core.log(f"tg vlm: result {len(desc)}c" if desc else "tg vlm: empty result")
    except Exception as e:
        core.log(f"tg vlm: {e}")
        desc = ""

    # Inject VLM description into main AI as [图片: desc] so AI can respond naturally
    user_msg = (f"[图片: {desc}]" + (f" {caption}" if caption else "")) if desc else (caption or "[图片，未识别]")

    def _do_reply() -> str:
        raw = core.get_reply(key, user_msg, prefer_minimax=True) or ""
        cleaned = strip_markers(raw, user_id=key)
        if cleaned:
            core.append_history(key, "user", f"[图片] {caption or ''}", user_id=key)
            core.append_history(key, "assistant", cleaned, user_id=key)
        return cleaned

    try:
        reply = await asyncio.to_thread(_do_reply)
    except Exception as e:
        core.log(f"tg photo reply: {e}")
        await msg.reply_text(f"(内部错误: {str(e)[:80]})")
        return

    if not reply:
        await msg.reply_text("(empty)")
        return

    bubbles = split_bubbles(reply) or [reply]
    await msg.reply_text(bubbles[0])
    if len(bubbles) > 1:
        await _send_bubbles(update, bubbles[1:])


async def on_voice(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    if await _reject_silent(update):
        return
    if update.message:
        await update.message.reply_text("暂不支持语音输入，发文字给我吧～")


async def handle_document(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    """Extract text from uploaded files and feed to AI (txt/pdf/docx/xlsx/code)."""
    if await _reject_silent(update):
        return
    doc = update.message.document
    if not doc:
        return
    try:
        await ctx.bot.send_chat_action(update.effective_chat.id, ChatAction.TYPING)
    except Exception:
        pass

    fname = (doc.file_name or "file").lower()
    ext = fname.rsplit(".", 1)[-1] if "." in fname else ""
    TEXT_EXTS = {
        "txt", "md", "py", "js", "ts", "json", "csv", "yaml", "yml",
        "xml", "html", "htm", "rst", "toml", "ini", "sh", "bat", "log", "sql",
    }

    try:
        tg_file = await ctx.bot.get_file(doc.file_id)
        data = bytes(await tg_file.download_as_bytearray())
    except Exception as e:
        core.log(f"handle_document: download error: {e}")
        await update.message.reply_text(f"文件下载失败: {e}")
        return

    def extract_text() -> str | None:
        import io as _io
        if ext in TEXT_EXTS:
            return data.decode("utf-8", errors="replace")
        if ext == "pdf":
            import pdfplumber
            with pdfplumber.open(_io.BytesIO(data)) as pdf:
                return "\n\n".join(p.extract_text() or "" for p in pdf.pages)
        if ext == "docx":
            import docx
            doc_obj = docx.Document(_io.BytesIO(data))
            return "\n".join(p.text for p in doc_obj.paragraphs if p.text.strip())
        if ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
            parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = [
                    "\t".join("" if c is None else str(c) for c in row)
                    for row in ws.iter_rows(values_only=True)
                    if any(c is not None for c in row)
                ]
                if rows:
                    parts.append(f"[Sheet: {sheet}]\n" + "\n".join(rows))
            return "\n\n".join(parts)
        return None

    try:
        text_content = await asyncio.to_thread(extract_text)
    except Exception as e:
        core.log(f"handle_document: extract error: {e}")
        await update.message.reply_text(f"文件读取失败: {e}")
        return

    if text_content is None:
        await update.message.reply_text(f"暂不支持 .{ext}，目前支持：txt/md/代码/pdf/docx/xlsx")
        return

    MAX_CHARS = 8000
    truncated = len(text_content) > MAX_CHARS
    if truncated:
        text_content = text_content[:MAX_CHARS]

    chat_id_int = update.effective_chat.id
    key = _chat_key(chat_id_int)
    caption = (update.message.caption or "").strip()
    instruction = f"（{caption}）" if caption else "请阅读并回应这份文件的内容。"
    user_msg = f"[文件: {doc.file_name}]\n{instruction}\n\n{text_content}"
    if truncated:
        user_msg += f"\n\n…（文件过长，只读了前 {MAX_CHARS} 字）"

    core.log(f"<< tg:{chat_id_int}: [doc] {doc.file_name} ext={ext} chars={len(text_content)}")

    def _do_reply() -> str:
        raw = core.get_reply(key, user_msg) or ""
        cleaned = strip_markers(raw, user_id=key)
        if cleaned:
            core.append_history(key, "user", f"[文件: {doc.file_name}] {caption or instruction}", user_id=key)
            core.append_history(key, "assistant", cleaned, user_id=key)
        return cleaned

    try:
        reply = await asyncio.to_thread(_do_reply)
    except Exception as e:
        core.log(f"tg handle_document reply: {e}")
        await update.message.reply_text(f"(内部错误: {str(e)[:80]})")
        return

    if not reply:
        await update.message.reply_text("(empty)")
        return

    bubbles = split_bubbles(reply) or [reply]
    await update.message.reply_text(bubbles[0])
    if len(bubbles) > 1:
        await _send_bubbles(update, bubbles[1:])


async def handle_location(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    """Reverse-geocode a shared location, save to memory, acknowledge."""
    if await _reject_silent(update):
        return
    loc = update.message.location
    if not loc:
        return
    lat, lon = loc.latitude, loc.longitude
    core.log(f"<< tg:{update.effective_chat.id}: [location] {lat},{lon}")

    def _geocode() -> str:
        try:
            return core.fetch_url_text(
                f"https://wttr.in/{lat},{lon}?format=%l|%C|%t|%w|%h&lang=zh", 400, 12
            )
        except Exception:
            return ""

    weather = await asyncio.to_thread(_geocode)
    place = weather.split("|", 1)[0].strip() if weather and "|" in weather else "?"

    key = _chat_key(update.effective_chat.id)
    memo = f"User location: {place} ({lat:.4f},{lon:.4f}), shared at {time.strftime('%Y-%m-%d %H:%M')}"
    await asyncio.to_thread(core.supa_bg, "memories", {"content": memo, "user_id": key})

    reply = f"位置收到了，记住了。{place}。"
    if weather:
        reply += f"\n天气：{weather}"
    await update.message.reply_text(reply)


async def _heartbeat(ctx: ContextTypes.DEFAULT_TYPE) -> None:
    HEARTBEAT_FILE.write_text(str(time.time()))


async def on_error(update: object, ctx: ContextTypes.DEFAULT_TYPE) -> None:
    core.log(f"tg error: {ctx.error}")
    # Fatal network errors that break the polling loop — trigger systemd restart
    err_str = str(ctx.error)
    if any(k in err_str for k in ("ConnectError", "RemoteProtocolError", "ConnectionResetError")):
        core.log("tg error: fatal network error, triggering restart via systemd")
        import os, signal
        os.kill(os.getpid(), signal.SIGTERM)


# ---------------------------------------------------------------------------
# Sync push API for cron scripts (no event loop needed)
# ---------------------------------------------------------------------------


def tg_push(text: str, chat_id: int | str | None = None) -> bool:
    """Send a single message via Telegram HTTP API. Returns True on success."""
    if not text:
        return False
    token = os.environ.get("TG_BOT_TOKEN") or TG_BOT_TOKEN
    if not token:
        core.log("tg_push: TG_BOT_TOKEN not set")
        return False
    target = str(chat_id or os.environ.get("TG_DEFAULT_CHAT_ID") or TG_DEFAULT_CHAT_ID or "")
    if not target:
        core.log("tg_push: no target chat_id")
        return False
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = urllib.parse.urlencode({"chat_id": target, "text": text}).encode()
    try:
        with urllib.request.urlopen(
            urllib.request.Request(url, data=payload, method="POST"), timeout=15
        ) as r:
            r.read()
        return True
    except Exception as e:
        core.log(f"tg_push: {e}")
        return False


def tg_push_bubbles(text: str, chat_id: int | str | None = None, delay: float = 0.5) -> int:
    """Split text into bubbles and push each one. Returns count sent."""
    if not text:
        return 0
    sent = 0
    for b in split_bubbles(text) or [text]:
        if tg_push(b, chat_id=chat_id):
            sent += 1
        time.sleep(delay)
    return sent


def tg_push_voice(audio_bytes: bytes, chat_id: int | str | None = None) -> bool:
    """Send a single voice message (audio bytes, mp3 ok) via Telegram sendVoice."""
    if not audio_bytes:
        return False
    token = os.environ.get("TG_BOT_TOKEN") or TG_BOT_TOKEN
    if not token:
        core.log("tg_push_voice: TG_BOT_TOKEN not set")
        return False
    target = str(chat_id or os.environ.get("TG_DEFAULT_CHAT_ID") or TG_DEFAULT_CHAT_ID or "")
    if not target:
        core.log("tg_push_voice: no target chat_id")
        return False
    url = f"https://api.telegram.org/bot{token}/sendVoice"
    boundary = f"----tgvoice{int(time.time()*1000)}"
    body = b"".join([
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"chat_id\"\r\n\r\n{target}\r\n".encode(),
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"voice\"; filename=\"v.mp3\"\r\n"
        "Content-Type: audio/mpeg\r\n\r\n".encode(),
        audio_bytes,
        f"\r\n--{boundary}--\r\n".encode(),
    ])
    try:
        req = urllib.request.Request(
            url, data=body, method="POST",
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        )
        with urllib.request.urlopen(req, timeout=30) as r:
            r.read()
        return True
    except Exception as e:
        core.log(f"tg_push_voice: {e}")
        return False


def tg_push_segments(text: str, chat_id: int | str | None = None,
                     user_id: str | None = None, record_history: bool = True) -> bool:
    """Cron-script push entry: parse [voice: ...] markers, send voice via TTS
    + sendVoice, send text via sendMessage, and append the cleaned text to
    history so the next conversation turn sees what was sent.

    Returns True if at least one segment was delivered.
    """
    if not text:
        return False
    target_chat = str(chat_id or os.environ.get("TG_DEFAULT_CHAT_ID") or TG_DEFAULT_CHAT_ID or "")
    history_key = f"tg_{target_chat}" if target_chat else None

    processed = core.process_markers(text, user_id=user_id or history_key)
    segs = core.parse_voice_segments(processed)
    if not segs:
        segs = [("text", processed)]

    sent_any = False
    for kind, content in segs:
        content = (content or "").strip()
        if not content:
            continue
        if kind == "voice":
            mp3 = core.generate_tts(content)
            if mp3 and tg_push_voice(mp3, chat_id=target_chat or None):
                core.log(f"tg_push_segments tts ok {len(mp3)}b")
                sent_any = True
            else:
                core.log(f"tg_push_segments tts fail, fallback text {len(content)}c")
                if tg_push(content, chat_id=target_chat or None):
                    sent_any = True
        else:
            cleaned_seg = _LEAK_PATTERN.sub("", content)
            cleaned_seg = re.sub(r"\n{3,}", "\n\n", cleaned_seg).strip()
            if cleaned_seg and tg_push_bubbles(cleaned_seg, chat_id=target_chat or None):
                sent_any = True

    if sent_any and record_history and history_key:
        cleaned_all = re.sub(r"\[voice:\s*(.*?)\]", r"\1", processed, flags=re.DOTALL)
        cleaned_all = _LEAK_PATTERN.sub("", cleaned_all)
        cleaned_all = re.sub(r"\n{3,}", "\n\n", cleaned_all).strip()
        if cleaned_all:
            try:
                core.append_history(history_key, "assistant", cleaned_all,
                                    user_id=user_id or history_key)
            except Exception as e:
                core.log(f"tg_push_segments append_history: {e}")
    return sent_any


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> None:
    if not TG_BOT_TOKEN:
        raise SystemExit("TG_BOT_TOKEN not set in environment")
    core.log(
        f"telegram bot starting "
        f"(whitelist={sorted(ALLOWED_CHAT_IDS) or 'OPEN'}, "
        f"default={TG_DEFAULT_CHAT_ID or 'unset'})"
    )
    builder = ApplicationBuilder().token(TG_BOT_TOKEN)
    if TG_PROXY:
        builder = builder.proxy(TG_PROXY).get_updates_proxy(TG_PROXY)
        core.log(f"tg using proxy: {TG_PROXY}")
    app = builder.build()
    HEARTBEAT_FILE.write_text(str(time.time()))  # initial heartbeat
    app.job_queue.run_repeating(_heartbeat, interval=60, first=10)
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("menu", cmd_start))
    app.add_handler(CommandHandler("help", cmd_start))
    app.add_handler(CommandHandler("clear", cmd_clear))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(CommandHandler("cost", cmd_cost))
    app.add_handler(CommandHandler("ds", cmd_ds))
    app.add_handler(CommandHandler("cf", cmd_cf))
    app.add_handler(CommandHandler("mm", cmd_mm))
    app.add_handler(CommandHandler("q", cmd_q))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.PHOTO, on_photo))
    app.add_handler(MessageHandler(filters.VOICE | filters.AUDIO, on_voice))
    app.add_handler(MessageHandler(filters.LOCATION, handle_location))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))
    app.add_error_handler(on_error)
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
